from django.db.models import manager
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import logout, login
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User, Group
from django.urls import reverse
from django.http import HttpResponseRedirect
from django.shortcuts import render, get_object_or_404, redirect
from allauth.account.signals import user_logged_in
from django.dispatch import receiver
from .forms import ClientRegistrationForm, CoachRegistrationForm, AthleteForm
from .models import Athlete
from .models import AthleteDocument
from .models import AthleteRequest
from .models import AthleteRoster
from .models import UserProfile
from .forms import UploadFileForm, UserEditForm
from .models import is_pma_admin
from django.core.files.storage import default_storage
import csv
import io
import requests
from openpyxl import load_workbook
from django.db.models import Q
from django.http import HttpResponse
from .forms import FeedbackFormSet
from .models import FeedbackRequest, FeedbackQuestion
from django.utils import timezone
from django.contrib import messages
import re

# Registration views for clients and coaches
def register_client(request):
    if request.method == 'POST':
        form = ClientRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('user_dashboard')  # Redirect to user dashboard after client registration
    else:
        form = ClientRegistrationForm()
    return render(request, 'registration/register.html', {'form': form})

def register_coach(request):
    if request.method == 'POST':
        form = CoachRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('coach_dashboard')  # Redirect to coach dashboard after coach registration
    else:
        form = CoachRegistrationForm()
    return render(request, 'registration/register.html', {'form': form})

@login_required
def choose_role(request):
    if request.method == 'POST':
        user_type = request.POST.get('user_type')

        # Redirect based on user choice
        if user_type == 'client':
            return redirect('user_dashboard')
        elif user_type == 'coach':
            return redirect('coach_dashboard')

    return render(request, 'choose_role.html')

@login_required
def coach_dashboard(request):
    # Fetch the logged-in coach's athletes
    roster = Athlete.objects.filter(coach=request.user)
    
    if request.method == 'POST':
        form = AthleteForm(request.POST, coach=request.user)
        if form.is_valid():
            if AthleteRequest.objects.filter(sender=request.user, receiver=form.cleaned_data.get('user')):
                messages.error(request, f"A request to {form.cleaned_data.get('user')} has already been sent!")
            else:
                form.save()
                messages.success(request, f"You've successfully sent a request to {form.cleaned_data.get('user')}") 
            return redirect('coach_dashboard')
    else:
        form = AthleteForm(coach=request.user)

    # Pass the athlete roster and the form to the template
    context = {
        'roster': roster,
        'form': form
    }
    
    return render(request, 'coach_dashboard.html', context)

@login_required
def user_dashboard(request):
    athlete = Athlete.objects.filter(user=request.user).first()
    documents = athlete.documents.all().order_by('-uploaded_at') if athlete else None
    parsed_data = []
    feedback_requests = FeedbackRequest.objects.filter(client=request.user).order_by('-created_at')
    
    # Check if feedback is already submitted
    feedback_submitted = all(question.answer for request in feedback_requests for question in request.questions.all())
     # Prepare notifications
    unread_feedback_requests = feedback_requests.filter(questions__answer__isnull=True).distinct()
    feedback_notification_count = 0 if feedback_submitted else unread_feedback_requests.count()
    quote = get_motivational_quote(request.user)

    # # Display a red alert if feedback has been submitted
    # if feedback_submitted:
    #     messages.error(request, "You have already submitted feedback. Please wait for a new request from your coach.")
    
    # Search functionality
    query = request.GET.get('query', '')  # Get the search term from the GET request
    if query and documents:
        documents = documents.filter(Q(keywords__icontains=query))

    # Check if there is a latest document and parse it if it's a CSV or XLSX
    if documents:
        latest_document = documents.first()  # Get the latest document

        try:
            response = requests.get(latest_document.document.url)
            response.raise_for_status()  # Ensure the request was successful

            if latest_document.is_csv:
                # Parse the CSV content
                file_content = io.StringIO(response.text)
                reader = csv.reader(file_content)
                parsed_data = [row for row in reader]
            elif latest_document.is_xlsx:
                # Parse the XLSX content
                file_content = io.BytesIO(response.content)
                workbook = load_workbook(file_content, data_only=True)
                sheet = workbook.active
                # Read rows from the active sheet
                parsed_data = [[cell.value for cell in row] for row in sheet.iter_rows()]
        except Exception as e:
            print(f"Error reading file: {e}")
    

    context = {
        'athlete': athlete,
        'documents': documents,
        'csv_data': parsed_data,  # Pass the parsed data to the template
        'query': query,  # Pass the query to the template for displaying in the search bar
        'feedback_requests': feedback_requests,
        'feedback_notification_count': feedback_notification_count,
        'quote': quote,
    }

    return render(request, 'user_dashboard.html', context)

# Logout view
def logout_view(request):
    logout(request)
    return redirect('home')

# Home view for logged out users or as a landing page
def home(request):
    return render(request, 'home.html')


def upload_form(request, athlete_id):
    athlete = get_object_or_404(Athlete, id=athlete_id)
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            # Save the new document as a separate entry without deleting any existing ones
            document = form.save(commit=False)
            document.athlete = athlete
            document.save()
            return redirect('coach_dashboard')
    else:
        form = UploadFileForm()

    return render(request, 'upload.html', {'form': form, 'athlete': athlete})

def extract_youtube_id(url):
    youtube_regex = r'(?:v=|\/v\/|embed\/|watch\?v=|youtu\.be\/|\/embed\/)([a-zA-Z0-9_-]{11})'
    match = re.search(youtube_regex, url)
    return match.group(1) if match else None

def view_document(request, athlete_id, document_id):
    athlete = get_object_or_404(Athlete, id=athlete_id)
    document = get_object_or_404(AthleteDocument, id=document_id, athlete=athlete)
    parsed_data = []

    try:
        response = requests.get(document.document.url)
        response.raise_for_status()

        if document.is_csv:
            # Parse CSV content
            file_content = io.StringIO(response.text)
            reader = csv.reader(file_content)
            parsed_data = [row for row in reader]
        elif document.is_xlsx:
            # Parse XLSX content
            file_content = io.BytesIO(response.content)
            workbook = load_workbook(file_content, data_only=True)
            sheet = workbook.active
            parsed_data = [[cell.value for cell in row] for row in sheet.iter_rows()]
    except Exception as e:
        print(f"Error reading file: {e}")

    # Add a flag to each video to check if it's a YouTube URL
    for video in document.videos.all():
        video.youtube_id = extract_youtube_id(video.url)

    context = {
        'athlete': athlete,
        'document': document,
        'parsed_data': parsed_data,
    }
    return render(request, 'view_document.html', context)
        
@login_required
def accept_athlete_request(request, request_id):
    athlete_request = get_object_or_404(AthleteRequest, id=request_id)
    existing_athlete = Athlete.objects.filter(coach=athlete_request.sender, user=athlete_request.receiver)
    existing_roster = AthleteRoster.objects.filter(manager=athlete_request.sender)
    if not existing_athlete:
        Athlete.objects.create(coach=athlete_request.sender, user=athlete_request.receiver, programming_deadline=athlete_request.programming_deadline)

        if not existing_roster:
           AthleteRoster.objects.create(manager=athlete_request.sender)

        roster = AthleteRoster.objects.get(manager=athlete_request.sender)
        roster.add_athlete(Athlete.objects.get(coach=athlete_request.sender, user=athlete_request.receiver, programming_deadline=athlete_request.programming_deadline))
        athlete_request.accept()
        messages.success(request, f"You've accepted {athlete_request.sender.username}'s request")
    else:
        athlete_request.decline()
        messages.info(request, "You are already a part of a coach's roster")
    return redirect('athlete_requests')
    

@login_required
def decline_athlete_request(request, request_id):
    athlete_request = get_object_or_404(AthleteRequest, id=request_id)
    athlete_request.decline()
    messages.success(request, f"You've declined {athlete_request.sender.username}'s request")
    return redirect('athlete_requests')


@login_required
def view_requests(request):
    requests = AthleteRequest.objects.filter(receiver = request.user)

    context = {
            'requests' : requests
            }

    return render(request, "athlete-requests.html", context)

@user_passes_test(is_pma_admin)
@login_required
def admin_dashboard(request):
    users = User.objects.all()
    # add admin functionalities here
    context = {
        'users': users,
        # add other context data as needed for admin dashboard
    }
    return render(request, 'admin_dashboard.html', context)

@login_required
def redirect_after_login(request):
    if is_pma_admin(request.user):
        return HttpResponseRedirect(reverse('admin_dashboard'))
    else:
        return HttpResponseRedirect(reverse('choose_role'))

@login_required
@user_passes_test(is_pma_admin)
def edit_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            return redirect('admin_dashboard')  # Redirect back to the admin dashboard after saving
    else:
        form = UserEditForm(instance=user)
    return render(request, 'edit_user.html', {'form': form, 'user': user})


@login_required
@user_passes_test(is_pma_admin)
def assign_role(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        role = request.POST.get('role')

        # Clear any existing roles to avoid duplicate role assignments
        user.groups.clear()

        if role == 'client':
            client_group, created = Group.objects.get_or_create(name='Client')
            user.groups.add(client_group)
        elif role == 'coach':
            coach_group, created = Group.objects.get_or_create(name='Coach')
            user.groups.add(coach_group)

        # Redirect back to the admin dashboard after assigning the role
        return redirect('admin_dashboard')

    # Render a form that lets the admin choose "Client" or "Coach" for this user
    return render(request, 'assign_role.html', {'user': user})

@login_required
@user_passes_test(is_pma_admin)
def change_role(request, user_id, group_id):
    user = get_object_or_404(User, id=user_id)
    group = get_object_or_404(Group, id=group_id)
    user.groups.clear()  # Clears current groups
    user.groups.add(group)  # Assigns the new group
    return redirect('admin_dashboard')

@login_required
@user_passes_test(is_pma_admin)
def remove_role(request, user_id, group_id):
    user = get_object_or_404(User, id=user_id)
    group = get_object_or_404(Group, id=group_id)
    user.groups.remove(group)
    return redirect('admin_dashboard')


@login_required
@user_passes_test(is_pma_admin)
def view_files(request, user_id):
    user = get_object_or_404(User, id=user_id)
    athlete = Athlete.objects.filter(user=user).first()
    documents = athlete.documents.all() if athlete else []

    context = {
        'user': user,
        'documents': documents,
    }
    return render(request, 'view_files.html', context)

@login_required
@user_passes_test(is_pma_admin)
def delete_file(request, document_id):
    document = get_object_or_404(AthleteDocument, id=document_id)
    document.document.delete()  # Deletes the file from storage
    document.delete()  # Deletes the record from the database
    return redirect('admin_dashboard')

@login_required
@user_passes_test(is_pma_admin)
def remove_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        # Confirm deletion
        user.delete()
        messages.success(request, f"User '{user.username}' has been successfully removed.")
        return redirect('admin_dashboard')

    # Render a confirmation page
    return render(request, 'remove_user.html', {'user': user})

@login_required
@user_passes_test(is_pma_admin)
def remove_file(request, file_id):
    file = get_object_or_404(AthleteDocument, id=file_id)
    user_id = file.athlete.user.id  # Get the user ID for redirection after deletion
    if request.method == 'POST':
        file.document.delete()  # This deletes the file from S3
        file.delete()  # Delete the file record from the database
        messages.success(request, "File removed successfully.")
    return redirect(reverse('view_files', args=[user_id]))

@login_required
def request_feedback(request):
    if request.method == 'POST':
        formset = FeedbackFormSet(request.POST)
        client_id = request.POST.get('client_id')
        client = get_object_or_404(User, id=client_id)

        if formset.is_valid():
            feedback_request = FeedbackRequest.objects.create(coach=request.user, client=client)
            for form in formset:
                if form.cleaned_data.get('question_text'):
                    FeedbackQuestion.objects.create(
                        feedback_request=feedback_request,
                        question_text=form.cleaned_data['question_text']
                    )
            messages.success(request, "Feedback request sent successfully!")
            return redirect('coach_dashboard')
    else:
        formset = FeedbackFormSet()
        clients = Athlete.objects.filter(coach=request.user).values_list('user__id', 'user__first_name')
        
    return render(request, 'request_feedback.html', {'formset': formset, 'clients': clients})

@login_required
def view_feedback_requests(request):
    feedback_requests = FeedbackRequest.objects.filter(client=request.user).order_by('-created_at')
    return render(request, 'feedback_requests.html', {'feedback_requests': feedback_requests})

@login_required
def submit_feedback(request, feedback_request_id):
    feedback_request = get_object_or_404(FeedbackRequest, id=feedback_request_id, client=request.user)
    if request.method == 'POST':
        # Save answers for each question
        for question in feedback_request.questions.all():
            answer = request.POST.get(f'answer_{question.id}')
            if answer:
                question.answer = answer
                question.save()

        # Send success message with coach's name
        messages.success(
            request, 
            f"Feedback submitted successfully to Coach {feedback_request.coach.first_name}!"
        )

        # Redirect to the user dashboard
        return redirect('user_dashboard')

    return render(request, 'feedback_requests.html', {'feedback_request': feedback_request})

@login_required
def read_feedback(request, athlete_id):
    athlete = get_object_or_404(Athlete, id=athlete_id, coach=request.user)
    feedback_requests = FeedbackRequest.objects.filter(client=athlete.user, coach=request.user)

    context = {
        'athlete': athlete,
        'feedback_requests': feedback_requests,
    }
    return render(request, 'read_feedback.html', context)

def faq(request):
    return render(request, 'faq.html')

def get_motivational_quote(user):

    user_profile, created = UserProfile.objects.get_or_create(user=user)

    if (user_profile.quote_timedate != timezone.now().date() and user_profile.motivational_quote) or not user_profile.motivational_quote:
        try:
            response = requests.get("https://quoteslate.vercel.app/api/quotes/random")
            data = response.json()
            user_profile.motivational_quote = data.get("quote")
            print("Motivational Quote Received:", user_profile.motivational_quote)
            user_profile.quote_timedate = timezone.now().date()
            user_profile.save()
        except:
            user_profile.motivational_quote = "Keep Moving Forward!"

    return user_profile.motivational_quote

def home(request):
    # Check if the logged-in user is part of the PMA Admin group
    is_admin = request.user.is_authenticated and request.user.groups.filter(name='PMA Admin').exists()
    return render(request, 'home.html', {
        'is_admin': is_admin,  # Pass the flag to the template
    })

def add_video(request, athlete_id, document_id):
    athlete = get_object_or_404(Athlete, id=athlete_id)
    document = get_object_or_404(AthleteDocument, id=document_id, athlete=athlete)

    if request.method == 'POST':
        title = request.POST.get('title')
        date_posted = request.POST.get('date_posted')
        url = request.POST.get('url')

        # Extract YouTube ID
        youtube_id = extract_youtube_id(url)
        if not youtube_id:
            messages.error(request, "Invalid YouTube URL. Please try again.")
            return HttpResponseRedirect(reverse('view_document', args=[athlete_id, document_id]))

        # Save the video information (assumes document.videos is set up in the model)
        document.videos.create(title=title, date_posted=date_posted, url=url, youtube_id=youtube_id)

        messages.success(request, "Video added successfully!")
        return HttpResponseRedirect(reverse('view_document', args=[athlete_id, document_id]))

    return render(request, 'view_document.html', {'athlete': athlete, 'document': document})

